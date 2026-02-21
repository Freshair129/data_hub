"""
Excel, Google Sheets & Facebook Ads Data Service
────────────────────────────────────────────────
Handles:
  1. Import data FROM Excel files
  2. Import data FROM Google Sheets
  3. Export data TO Excel WITH formulas
  4. Export data TO Google Sheets
  5. Fetch Ad Data FROM Facebook Graph API

Dependencies: openpyxl, xlsxwriter, gspread, google-auth, facebook-business
"""

import os
import json
import sys
import uuid # For generating CUID/UUIDs
from datetime import datetime, date
from db_adapter import get_db_conn # Import for SQL access

# ─── Check Optional Dependencies ───────────────────────────
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlsxwriter
    HAS_XLSXWRITER = True
except ImportError:
    HAS_XLSXWRITER = False

try:
    import gspread
    from google.oauth2.service_account import Credentials
    HAS_GSPREAD = True
except ImportError:
    HAS_GSPREAD = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    from facebook_business.api import FacebookAdsApi
    from facebook_business.adobjects.adaccount import AdAccount
    from facebook_business.adobjects.campaign import Campaign
    from facebook_business.adobjects.adset import AdSet
    from facebook_business.adobjects.ad import Ad
    from facebook_business.adobjects.adcreative import AdCreative
    HAS_FB_SDK = True
except ImportError:
    HAS_FB_SDK = False


# ═══════════════════════════════════════════════════════════
#  IMPORT: Read from Excel
# ═══════════════════════════════════════════════════════════

def import_from_excel(file_path, sheet_name=None):
    """
    Read an Excel file and return data as list of dicts.
    Supports .xlsx and .xls formats.
    """
    if HAS_PANDAS:
        df = pd.read_excel(file_path, sheet_name=sheet_name or 0)
        return df.to_dict(orient='records')
    
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required: pip install openpyxl")
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    
    headers = [cell.value for cell in ws[1]]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {}
        for i, value in enumerate(row):
            if i < len(headers) and headers[i]:
                record[headers[i]] = value
        if any(v is not None for v in record.values()):
            data.append(record)
    
    wb.close()
    return data


# ═══════════════════════════════════════════════════════════
#  IMPORT: Read from Google Sheets
# ═══════════════════════════════════════════════════════════

def import_from_google_sheets(spreadsheet_id, sheet_name=None, credentials_path=None):
    """
    Read data from a Google Sheet.
    Requires a service account JSON credentials file.
    """
    if not HAS_GSPREAD:
        raise ImportError("gspread is required: pip install gspread google-auth")
    
    creds_path = credentials_path or os.getenv('GOOGLE_SHEETS_CREDENTIALS')
    if not creds_path:
        raise ValueError("Google Sheets credentials not configured. Set GOOGLE_SHEETS_CREDENTIALS env var.")
    
    scopes = [
        'https://www.googleapis.com/auth/spreadsheets',
        'https://www.googleapis.com/auth/drive'
    ]
    creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
    client = gspread.authorize(creds)
    
    spreadsheet = client.open_by_key(spreadsheet_id)
    worksheet = spreadsheet.worksheet(sheet_name) if sheet_name else spreadsheet.sheet1
    
    return worksheet.get_all_records()


# ═══════════════════════════════════════════════════════════
#  EXPORT: Write to Excel WITH Formulas
# ═══════════════════════════════════════════════════════════

def export_customers_to_excel(customers, output_path, include_formulas=True):
    """
    Export customer data to Excel with rich formatting and formulas.
    
    Formulas included:
      - Total Revenue (SUM)
      - Average Order Value (AVERAGE)
      - Customer Count (COUNTA)
      - Conversion Rate (COUNTIF)
    """
    if not HAS_XLSXWRITER:
        raise ImportError("xlsxwriter is required: pip install xlsxwriter")
    
    wb = xlsxwriter.Workbook(output_path)
    
    # ─── Formats ────────────────────────────────────────────
    header_fmt = wb.add_format({
        'bold': True, 'font_size': 12,
        'bg_color': '#1a1a2e', 'font_color': '#e94560',
        'border': 1, 'text_wrap': True, 'valign': 'vcenter'
    })
    currency_fmt = wb.add_format({'num_format': '#,##0.00 "฿"', 'border': 1})
    date_fmt = wb.add_format({'num_format': 'yyyy-mm-dd', 'border': 1})
    cell_fmt = wb.add_format({'border': 1, 'text_wrap': True, 'valign': 'top'})
    title_fmt = wb.add_format({
        'bold': True, 'font_size': 16,
        'font_color': '#e94560', 'bottom': 2
    })
    summary_label_fmt = wb.add_format({'bold': True, 'bg_color': '#f0f0f0', 'border': 1})
    summary_value_fmt = wb.add_format({
        'bold': True, 'num_format': '#,##0.00 "฿"',
        'bg_color': '#f0f0f0', 'border': 1, 'font_color': '#16213e'
    })
    percent_fmt = wb.add_format({
        'bold': True, 'num_format': '0.0%',
        'bg_color': '#f0f0f0', 'border': 1, 'font_color': '#0f3460'
    })
    score_fmt = wb.add_format({'num_format': '0', 'border': 1})
    
    # ═══ Sheet 1: Customer Master List ═════════════════════
    ws1 = wb.add_worksheet('Customers')
    ws1.set_tab_color('#e94560')
    
    headers = [
        'Customer ID', 'Name', 'Phone', 'Email', 'Status',
        'Tier', 'Stage', 'Join Date', 'Total Spend',
        'Total Orders', 'AI Score', 'AI Intent', 'Tags'
    ]
    
    # Title Row
    ws1.merge_range('A1:M1', f'V-School CRM — Customer Report ({date.today().isoformat()})', title_fmt)
    
    # Header Row
    for col, header in enumerate(headers):
        ws1.write(2, col, header, header_fmt)
    
    # Data Rows
    row = 3
    for cust in customers:
        profile = cust.get('profile', {})
        contact = cust.get('contact_info', {})
        intel = cust.get('intelligence', {})
        metrics = intel.get('metrics', {})
        
        name = f"{profile.get('first_name', '')} {profile.get('last_name', '')}".strip()
        tags = ', '.join(intel.get('tags', []))
        
        ws1.write(row, 0, cust.get('customer_id', ''), cell_fmt)
        ws1.write(row, 1, name, cell_fmt)
        ws1.write(row, 2, contact.get('phone_primary', ''), cell_fmt)
        ws1.write(row, 3, contact.get('email', ''), cell_fmt)
        ws1.write(row, 4, profile.get('status', ''), cell_fmt)
        ws1.write(row, 5, profile.get('membership_tier', ''), cell_fmt)
        ws1.write(row, 6, profile.get('lifecycle_stage', ''), cell_fmt)
        
        join_date_str = profile.get('join_date', '')
        if join_date_str:
            try:
                ws1.write_datetime(row, 7, datetime.strptime(join_date_str, '%Y-%m-%d'), date_fmt)
            except:
                ws1.write(row, 7, join_date_str, cell_fmt)
        else:
            ws1.write(row, 7, '', cell_fmt)
        
        ws1.write_number(row, 8, metrics.get('total_spend', 0), currency_fmt)
        ws1.write_number(row, 9, metrics.get('total_order', 0), cell_fmt)
        ws1.write_number(row, 10, intel.get('score', 0), score_fmt)
        ws1.write(row, 11, intel.get('intent', ''), cell_fmt)
        ws1.write(row, 12, tags, cell_fmt)
        row += 1
    
    last_data_row = row
    
    # ─── Summary Formulas ───────────────────────────────────
    if include_formulas and len(customers) > 0:
        summary_row = last_data_row + 2
        
        ws1.write(summary_row, 0, 'SUMMARY', summary_label_fmt)
        ws1.merge_range(summary_row, 1, summary_row, 2, '', summary_label_fmt)
        
        # Total Customers
        ws1.write(summary_row + 1, 0, 'Total Customers', summary_label_fmt)
        ws1.write_formula(summary_row + 1, 1,
            f'=COUNTA(A4:A{last_data_row})', cell_fmt, len(customers))
        
        # Total Revenue
        ws1.write(summary_row + 2, 0, 'Total Revenue', summary_label_fmt)
        ws1.write_formula(summary_row + 2, 1,
            f'=SUM(I4:I{last_data_row})', summary_value_fmt)
        
        # Average Order Value
        ws1.write(summary_row + 3, 0, 'Avg Order Value', summary_label_fmt)
        ws1.write_formula(summary_row + 3, 1,
            f'=IF(SUM(J4:J{last_data_row})>0, SUM(I4:I{last_data_row})/SUM(J4:J{last_data_row}), 0)',
            summary_value_fmt)
        
        # Average AI Score
        ws1.write(summary_row + 4, 0, 'Avg AI Score', summary_label_fmt)
        ws1.write_formula(summary_row + 4, 1,
            f'=AVERAGE(K4:K{last_data_row})', cell_fmt)
        
        # Conversion Rate (Paid / Total)
        ws1.write(summary_row + 5, 0, 'Conversion Rate', summary_label_fmt)
        ws1.write_formula(summary_row + 5, 1,
            f'=IF(COUNTA(E4:E{last_data_row})>0,'
            f'COUNTIF(E4:E{last_data_row},"Active")/COUNTA(E4:E{last_data_row}), 0)',
            percent_fmt)
        
        # Tier Breakdown
        ws1.write(summary_row + 7, 0, 'TIER BREAKDOWN', summary_label_fmt)
        for i, tier in enumerate(['MEMBER', 'SILVER', 'GOLD', 'VIP']):
            ws1.write(summary_row + 8 + i, 0, tier, cell_fmt)
            ws1.write_formula(summary_row + 8 + i, 1,
                f'=COUNTIF(F4:F{last_data_row},"{tier}")', cell_fmt)
    
    # Column widths
    col_widths = [14, 20, 15, 25, 10, 10, 10, 12, 14, 10, 10, 12, 25]
    for i, w in enumerate(col_widths):
        ws1.set_column(i, i, w)
    
    # ═══ Sheet 2: Orders Detail ════════════════════════════
    ws2 = wb.add_worksheet('Orders')
    ws2.set_tab_color('#0f3460')
    
    order_headers = ['Order ID', 'Customer ID', 'Customer Name', 'Date', 'Status', 'Amount', 'Paid']
    
    ws2.merge_range('A1:G1', 'Orders Detail', title_fmt)
    for col, h in enumerate(order_headers):
        ws2.write(2, col, h, header_fmt)
    
    orow = 3
    for cust in customers:
        profile = cust.get('profile', {})
        name = f"{profile.get('first_name', '')} {profile.get('last_name', '')}".strip()
        for order in cust.get('orders', []):
            ws2.write(orow, 0, order.get('order_id', ''), cell_fmt)
            ws2.write(orow, 1, cust.get('customer_id', ''), cell_fmt)
            ws2.write(orow, 2, name, cell_fmt)
            ws2.write(orow, 3, order.get('date', ''), date_fmt)
            ws2.write(orow, 4, order.get('status', ''), cell_fmt)
            ws2.write_number(orow, 5, order.get('total_amount', 0), currency_fmt)
            ws2.write_number(orow, 6, order.get('paid_amount', 0), currency_fmt)
            orow += 1
    
    if orow > 3:
        ws2.write(orow + 1, 4, 'TOTAL', summary_label_fmt)
        ws2.write_formula(orow + 1, 5, f'=SUM(F4:F{orow})', summary_value_fmt)
        ws2.write_formula(orow + 1, 6, f'=SUM(G4:G{orow})', summary_value_fmt)
    
    ws2.set_column(0, 0, 16)
    ws2.set_column(1, 1, 14)
    ws2.set_column(2, 2, 20)
    ws2.set_column(3, 3, 12)
    ws2.set_column(4, 4, 10)
    ws2.set_column(5, 6, 14)
    
    # ═══ Sheet 3: AI Intelligence ═════════════════════════
    ws3 = wb.add_worksheet('AI Intelligence')
    ws3.set_tab_color('#e94560')
    
    ai_headers = ['Customer', 'Lead Score', 'Intent', 'Interest', 'Churn Risk', 'Last AI Update']
    ws3.merge_range('A1:F1', 'AI Intelligence Report', title_fmt)
    for col, h in enumerate(ai_headers):
        ws3.write(2, col, h, header_fmt)
    
    arow = 3
    for cust in customers:
        profile = cust.get('profile', {})
        intel = cust.get('intelligence', {})
        name = f"{profile.get('first_name', '')} {profile.get('last_name', '')}".strip()
        
        ws3.write(arow, 0, name, cell_fmt)
        ws3.write_number(arow, 1, intel.get('score', 0), score_fmt)
        ws3.write(arow, 2, intel.get('intent', 'Unknown'), cell_fmt)
        ws3.write(arow, 3, intel.get('main_interest', ''), cell_fmt)
        ws3.write(arow, 4, intel.get('churn_risk', 'N/A'), cell_fmt)
        ws3.write(arow, 5, intel.get('last_ai_update', ''), cell_fmt)
        arow += 1
    
    ws3.set_column(0, 0, 20)
    ws3.set_column(1, 1, 12)
    ws3.set_column(2, 5, 15)
    
    wb.close()
    print(f"[Excel] ✅ Exported {len(customers)} customers to: {output_path}")
    return output_path


# ═══════════════════════════════════════════════════════════
#  EXPORT: Push to Google Sheets
# ═══════════════════════════════════════════════════════════

def export_to_google_sheets(data, spreadsheet_id, sheet_name='CRM Data', credentials_path=None):
    """
    Write data to a Google Sheet.
    `data` should be a list of dicts.
    """
    if not HAS_GSPREAD:
        raise ImportError("gspread is required")
    
    creds_path = credentials_path or os.getenv('GOOGLE_SHEETS_CREDENTIALS')
    scopes = [
        'https://www.googleapis.com/auth/spreadsheets',
        'https://www.googleapis.com/auth/drive'
    ]
    creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
    client = gspread.authorize(creds)
    
    spreadsheet = client.open_by_key(spreadsheet_id)
    
    try:
        worksheet = spreadsheet.worksheet(sheet_name)
        worksheet.clear()
    except gspread.exceptions.WorksheetNotFound:
        worksheet = spreadsheet.add_worksheet(title=sheet_name, rows=len(data)+10, cols=20)
    
    if not data:
        return
    
    headers = list(data[0].keys())
    rows = [headers]
    for item in data:
        rows.append([serialize_value(item.get(h, '')) for h in headers])
    
    worksheet.update(range_name='A1', values=rows)
    print(f"[Google Sheets] ✅ Exported {len(data)} rows to sheet '{sheet_name}'")


def serialize_value(val):
    """Ensure all values are JSON-serializable for Google Sheets."""
    if val is None:
        return ''
    if isinstance(val, (datetime, date)):
        return val.isoformat()
    if isinstance(val, (dict, list)):
        return json.dumps(val, ensure_ascii=False)
    return val


# ═══════════════════════════════════════════════════════════
#  CLI Entry Point
# ═══════════════════════════════════════════════════════════

def main():
    """
    Usage:
      python data_service.py export-excel [output_path]
      python data_service.py import-excel <file_path>
      python data_service.py import-gsheets <spreadsheet_id> [sheet_name]
      python data_service.py export-gsheets <spreadsheet_id> [sheet_name]
      python data_service.py fetch-ads <ad_account_id> [output.json]
    """
    if len(sys.argv) < 2:
        print(__doc__)
        return
    
    command = sys.argv[1]
    
    if command == 'export-excel':
        output = sys.argv[2] if len(sys.argv) > 2 else f'crm_export_{date.today().isoformat()}.xlsx'
        customers = load_all_customers_json()
        export_customers_to_excel(customers, output)
    
    elif command == 'import-excel':
        file_path = sys.argv[2]
        data = import_from_excel(file_path)
        print(json.dumps(data[:5], indent=2, ensure_ascii=False, default=str))
        print(f"... ({len(data)} total records)")
    
    elif command == 'import-gsheets':
        sheet_id = sys.argv[2]
        sheet_name = sys.argv[3] if len(sys.argv) > 3 else None
        data = import_from_google_sheets(sheet_id, sheet_name)
        print(json.dumps(data[:5], indent=2, ensure_ascii=False, default=str))
    
    elif command == 'export-gsheets':
        sheet_id = sys.argv[2]
        sheet_name = sys.argv[3] if len(sys.argv) > 3 else 'CRM Data'
        customers = load_all_customers_json()
        flat = [flatten_customer(c) for c in customers]
        export_to_google_sheets(flat, sheet_id, sheet_name)
    
    elif command == 'fetch-ads':
        ad_account_id = sys.argv[2]
        output_file = sys.argv[3] if len(sys.argv) > 3 else f"ads_data_{date.today().isoformat()}.json"
        
        data = fetch_facebook_ads(ad_account_id)
        
        if data:
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            print(f"💾 Saved Ad Data to: {output_file}")
            
            # Save to Database
            save_ad_data_to_db(data)
    
    else:
        print(f"Unknown command: {command}")


# ═══════════════════════════════════════════════════════════
#  FACEBOOK ADS DATA FETCHING
# ═══════════════════════════════════════════════════════════

def fetch_facebook_ads(ad_account_id):
    """
    Fetch Campaigns, AdSets, Ads, and Creatives from Facebook Marketing API.
    Returns a dictionary with the data structure.
    """
    if not HAS_FB_SDK:
        raise ImportError("facebook-business is required. pip install facebook-business")

    # 1. Configuration
    access_token = os.getenv('FB_PAGE_ACCESS_TOKEN') # Using Page Token for simplicity (ensure permissions)
    app_id = os.getenv('FB_APP_ID')
    app_secret = os.getenv('FB_APP_SECRET')

    if not access_token:
        print("[Error] FB_PAGE_ACCESS_TOKEN is missing.")
        return None
    
    # 2. Initialize API
    try:
        FacebookAdsApi.init(access_token=access_token)
    except Exception as e:
        print(f"[Error] Failed to init Facebook API: {e}")
        return None

    # Ensure account ID starts with 'act_'
    account_id = f"act_{ad_account_id}" if not ad_account_id.startswith("act_") else ad_account_id
    account = AdAccount(account_id)
    
    data = {
        "campaigns": [],
        "adsets": [],
        "ads": [],
        "creatives": []
    }

    try:
        print(f"🔄 Fetching data for Ad Account: {account_id}...")

        # 3. Fetch Campaigns
        campaign_fields = [
            Campaign.Field.id,
            Campaign.Field.name,
            Campaign.Field.status,
            Campaign.Field.objective,
            Campaign.Field.start_time,
            Campaign.Field.stop_time,
        ]
        campaigns = account.get_campaigns(fields=campaign_fields)
        data["campaigns"] = [c.export_all_data() for c in campaigns]
        print(f"✅ Found {len(data['campaigns'])} Campaigns")

        # 4. Fetch AdSets
        adset_fields = [
            AdSet.Field.id,
            AdSet.Field.name,
            AdSet.Field.status,
            AdSet.Field.daily_budget,
            AdSet.Field.campaign_id,
            AdSet.Field.targeting,
        ]
        adsets = account.get_ad_sets(fields=adset_fields)
        data["adsets"] = [a.export_all_data() for a in adsets]
        print(f"✅ Found {len(data['adsets'])} AdSets")

        # 5. Fetch Ads
        ad_fields = [
            Ad.Field.id,
            Ad.Field.name,
            Ad.Field.status,
            Ad.Field.adset_id,
            Ad.Field.creative,
        ]
        ads = account.get_ads(fields=ad_fields)
        data["ads"] = [a.export_all_data() for a in ads]
        print(f"✅ Found {len(data['ads'])} Ads")

        # 6. Fetch Creatives
        creative_fields = [
            AdCreative.Field.id,
            AdCreative.Field.name,
            AdCreative.Field.body,
            AdCreative.Field.title,
            AdCreative.Field.image_url,
            AdCreative.Field.thumbnail_url,
            AdCreative.Field.call_to_action_type,
        ]
        # Fetching creatives from the account to ensure we get all of them, or could iterate ads.
        # Account level fetch is better for bulk.
        creatives = account.get_ad_creatives(fields=creative_fields, params={'limit': 100})
        data["creatives"] = [c.export_all_data() for c in creatives]
        print(f"✅ Found {len(data['creatives'])} Creatives")

        return data

    except Exception as e:
        print(f"[Error] Failed to fetch Facebook Ads data: {e}")
        return None


def save_ad_data_to_db(data):
    """
    Saves the fetched Ad Data dict to PostgreSQL using db_adapter connection.
    Upserts AdAccount, Campaign, AdSet, AdCreative, Ad.
    """
    if not data: return False
    
    conn = get_db_conn()
    if not conn:
        print("[Error] No DB Connection available (DB_ADAPTER != prisma?)")
        return False
        
    try:
        cur = conn.cursor()
        
        # 1. UPSERT AdAccount (We assume current account for now, or extract from data if available)
        # For this script, we passed ad_account_id to fetch, we should probably pass it here too or just handle it.
        # Ideally, we upsert campaigns first and link them.
        
        # 2. UPSERT Campaigns
        # Schema: id, name, status, objective, startDate, endDate, createdAt, updatedAt
        print(f"💾 Saving {len(data['campaigns'])} Campaigns...")
        for c in data['campaigns']:
            cur.execute("""
                INSERT INTO campaigns (id, name, status, objective, start_date, created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, NOW(), NOW())
                ON CONFLICT (id) DO UPDATE SET
                    name = EXCLUDED.name,
                    status = EXCLUDED.status,
                    objective = EXCLUDED.objective,
                    start_date = EXCLUDED.start_date,
                    updated_at = NOW();
            """, (
                c.get('id'),
                c.get('name'),
                c.get('status'),
                c.get('objective'),
                c.get('start_time') # FB returns ISO string or None
            ))
            
        # 3. UPSERT AdSets
        # Schema: id, adSetId, name, status, campaignId, dailyBudget, targeting
        print(f"💾 Saving {len(data['adsets'])} AdSets...")
        for a in data['adsets']:
            # Try to find existing by adSetId
            cur.execute("SELECT id FROM ad_sets WHERE ad_set_id = %s", (a.get('id'),))
            res = cur.fetchone()
            
            if res:
                # Update
                cur.execute("""
                    UPDATE ad_sets SET 
                        name = %s, status = %s, daily_budget = %s, targeting = %s, updated_at = NOW()
                    WHERE ad_set_id = %s
                """, (
                    a.get('name'), a.get('status'), int(a.get('daily_budget') or 0)/100, 
                    json.dumps(a.get('targeting') or {}), a.get('id')
                ))
            else:
                # Insert
                new_id = f"c{uuid.uuid4().hex}" # Pseudo-CUID
                cur.execute("""
                    INSERT INTO ad_sets (id, ad_set_id, campaign_id, name, status, daily_budget, targeting, created_at, updated_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                """, (
                    new_id, a.get('id'), a.get('campaign_id'), 
                    a.get('name'), a.get('status'), int(a.get('daily_budget') or 0)/100,
                    json.dumps(a.get('targeting') or {})
                ))

        # 4. UPSERT Creatives
        # Schema: id, name, body, headline, imageUrl, videoUrl, callToAction
        print(f"💾 Saving {len(data['creatives'])} Creatives...")
        for c in data['creatives']:
            cur.execute("""
                INSERT INTO ad_creatives (id, name, body, headline, image_url, video_url, call_to_action, created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                ON CONFLICT (id) DO UPDATE SET
                    name = EXCLUDED.name,
                    body = EXCLUDED.body,
                    headline = EXCLUDED.headline,
                    image_url = EXCLUDED.image_url,
                    video_url = EXCLUDED.video_url,
                    call_to_action = EXCLUDED.call_to_action,
                    updated_at = NOW();
            """, (
                c.get('id'), # FB Creative ID as Primary Key
                c.get('name'),
                c.get('body'),
                c.get('title'),
                c.get('image_url') or c.get('thumbnail_url'),
                c.get('video_url'), 
                c.get('call_to_action_type')
            ))

        # 5. UPSERT Ads
        # Schema: id, adId, name, status, adSetId, creativeId, spend, etc.
        print(f"💾 Saving {len(data['ads'])} Ads...")
        for ad in data['ads']:
            # Lookup AdSet CUID
            cur.execute("SELECT id FROM ad_sets WHERE ad_set_id = %s", (ad.get('adset_id'),))
            res_set = cur.fetchone()
            if not res_set:
                print(f"[Warn] AdSet {ad.get('adset_id')} not found for Ad {ad.get('name')}")
                continue
            real_adset_id = res_set[0]
            
            # Lookup Creative ID
            creative_fb_id = ad.get('creative', {}).get('id')
            
            # Upsert Ad
            cur.execute("SELECT id FROM ads WHERE ad_id = %s", (ad.get('id'),))
            res_ad = cur.fetchone()
            
            if res_ad:
                cur.execute("""
                    UPDATE ads SET 
                        name = %s, status = %s, ad_set_id = %s, creative_id = %s, updated_at = NOW()
                    WHERE ad_id = %s
                """, (
                    ad.get('name'), ad.get('status'), real_adset_id, creative_fb_id, ad.get('id')
                ))
            else:
                new_ad_uuid = f"ad{uuid.uuid4().hex}"
                cur.execute("""
                    INSERT INTO ads (id, ad_id, name, status, ad_set_id, creative_id, created_at, updated_at)
                    VALUES (%s, %s, %s, %s, %s, %s, NOW(), NOW())
                """, (
                    new_ad_uuid, ad.get('id'), ad.get('name'), ad.get('status'), 
                    real_adset_id, creative_fb_id
                ))
        
        print("✅ Database Population Complete!")

    except Exception as e:
        print(f"[DB/Error] {e}")
        return False

    return True



def load_all_customers_json():
    """Load all customers from local JSON files."""
    data_dir = os.path.join(os.getcwd(), '..', 'customer')
    customers = []
    
    if not os.path.exists(data_dir):
        print(f"[Warning] Customer directory not found: {data_dir}")
        return customers
    
    for folder in os.listdir(data_dir):
        folder_path = os.path.join(data_dir, folder)
        if not os.path.isdir(folder_path) or folder.startswith('.'):
            continue
        
        for f in os.listdir(folder_path):
            if f.startswith('profile_') and f.endswith('.json'):
                try:
                    with open(os.path.join(folder_path, f), 'r', encoding='utf-8') as fh:
                        customers.append(json.load(fh))
                except Exception as e:
                    print(f"[Warning] Failed to read {f}: {e}")
    
    return customers


def flatten_customer(cust):
    """Flatten nested customer JSON to a flat dict for Google Sheets."""
    profile = cust.get('profile', {})
    contact = cust.get('contact_info', {})
    intel = cust.get('intelligence', {})
    metrics = intel.get('metrics', {})
    
    return {
        'customer_id': cust.get('customer_id', ''),
        'name': f"{profile.get('first_name', '')} {profile.get('last_name', '')}".strip(),
        'phone': contact.get('phone_primary', ''),
        'email': contact.get('email', ''),
        'status': profile.get('status', ''),
        'tier': profile.get('membership_tier', ''),
        'stage': profile.get('lifecycle_stage', ''),
        'join_date': profile.get('join_date', ''),
        'total_spend': metrics.get('total_spend', 0),
        'total_orders': metrics.get('total_order', 0),
        'ai_score': intel.get('score', 0),
        'ai_intent': intel.get('intent', ''),
        'tags': ', '.join(intel.get('tags', []))
    }


if __name__ == '__main__':
    main()
